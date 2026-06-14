import openpyxl

template_path = r"c:\Users\joelz\Dokumente\AeroTrackTiming\Rangliste_Vorlage.xlsx"
out_path = r"c:\Users\joelz\Dokumente\AeroTrackTiming\scratch\deep_excel_structure.txt"

wb = openpyxl.load_workbook(template_path)

with open(out_path, "w", encoding="utf-8") as f:
    for name in wb.sheetnames:
        ws = wb[name]
        f.write(f"\nSheet: {name}\n")
        f.write(f"Max row: {ws.max_row}, Max col: {ws.max_column}\n")
        f.write(f"Merged ranges: {ws.merged_cells.ranges}\n")
        # Print row heights
        for r in range(1, ws.max_row + 1):
            h = ws.row_dimensions[r].height
            if h is not None:
                f.write(f"Row {r} height: {h}\n")
        # Print col widths
        for c in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            w = ws.column_dimensions[col_letter].width
            if w is not None:
                f.write(f"Col {col_letter} width: {w}\n")
                
        # Print non-empty cells
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value is not None:
                    border_info = ""
                    if cell.border:
                        b_parts = []
                        if cell.border.left and cell.border.left.style: b_parts.append(f"L={cell.border.left.style}")
                        if cell.border.right and cell.border.right.style: b_parts.append(f"R={cell.border.right.style}")
                        if cell.border.top and cell.border.top.style: b_parts.append(f"T={cell.border.top.style}")
                        if cell.border.bottom and cell.border.bottom.style: b_parts.append(f"B={cell.border.bottom.style}")
                        if b_parts: border_info = " border:(" + ",".join(b_parts) + ")"
                    
                    fill_info = ""
                    if cell.fill and cell.fill.fill_type:
                        fill_info = f" fill:({cell.fill.fill_type})"
                        if cell.fill.start_color and cell.fill.start_color.rgb:
                            fill_info += f" color:({cell.fill.start_color.rgb})"
                    
                    font_info = ""
                    if cell.font:
                        font_info = f" font:({cell.font.name},{cell.font.size},{'bold' if cell.font.bold else 'regular'})"
                    
                    f.write(f"Cell {cell.coordinate}: value={repr(cell.value)}{font_info}{fill_info}{border_info}\n")
print("Done!")
